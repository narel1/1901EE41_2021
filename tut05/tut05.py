import openpyxl
import os.path
import csv

def generate_marksheetname(name_dict, sub_name, LTP_dict, student_info,  score_dict, roll_no):
    wb = openpyxl.Workbook()
    sheetname = wb.active
    sheetname.title="Overall"           #create first sheetname for overall performance
                               
    #prepare first_line for the sheetname to store each semester's score
    first_line=['Sl No.','Subject No.','Subject Name','L-T-P','Credit','Subject Type','Grade']
    last_sem=0       #stores value of last semester
    sl_no=1          #stores serial number
    id_no=1          #stores the index number of the sheetname
    
    for row in student_info:
        sheetnamename="Sem"+row[1]     #stores the name of sheetname
        if int(row[1])!=last_sem:  #current semester is not same as last semester
            sl_no=1                  #again make the serial number as 1
            last_sem=int(row[1])   #change value of last semester to current semester
            wb.create_sheetname(index=id_no, title=sheetnamename) #create sheetname for new semester
            wb[sheetnamename].append(first_line)                    #add first_line to the new sheetname
            id_no=id_no+1                              #increase index number by 1

        #create a list named grade with required student_info and append it to the sheetname
        grade=[sl_no, row[2], sub_name[row[2]], LTP_dict[row[2]], row[3], row[5], row[4]]
        wb[sheetnamename].append(grade)
        sl_no=sl_no+1                 #increase serial number by 1

    #give all the headings for the first column
    sheetname['A1']="Roll No."          
    sheetname['B1']=roll_no           #fill roll number of the student    
    sheetname['A2']="Name of Student"   
    sheetname.merge_cells('B2:C2')      #combine cell B2 and C2
    sheetname.cell(row = 2, column = 2).value = name_dict[roll_no]  #fill student's name with that roll_no
    sheetname['A3']="Discipline"        
    Branch=roll_no[4:6]       #get name of branch
    sheetname['B3']=Branch          #fill value of branch
    sheetname['A4']="Semester No."      
    sheetname['A5']="Semester wise Credit Taken"    
    sheetname['A6']="SPI"                           
    sheetname['A7']="Total Credits Taken"           
    sheetname['A8']="CPI"

    names_of_sheetname=wb.sheetnamenames                  #get a list of sheetnamenames
    total_credit=0                            #stores total number of credits taken
    total_obtained=0 

    for i in range(1,len(names_of_sheetname)):
        sem_sheetname=wb[names_of_sheetname[i]]           #sheetname for corresponding semester
        sem_no=names_of_sheetname[i][3:]              #get semester number from sheetnamename
        column_letter=openpyxl.utils.get_column_letter(i+1) #get column letter from number
        sheetname[column_letter+str(4)]=int(sem_no)             #fill semester number in 4th line
        semwise_credit=0                      #stores total credit per semester 
        semwise_obtained=0                    
        row_count = sem_sheetname.max_row         #total number of lines in sheetname for that semester

        for j in range(2, row_count + 1): 
            credit=int(sem_sheetname['E'+str(j)].value)  #get credit of particular course
            grade=sem_sheetname['G'+str(j)].value        #grade obtained in that course
            semwise_credit=semwise_credit+credit     #update total credit in a given semester
            semwise_obtained=semwise_obtained+(credit*score_dict[grade])  # apply formula
        total_credit=total_credit+semwise_credit     #total credit till the given semester
        total_obtained=total_obtained+semwise_obtained
        SPI_obtained=semwise_obtained/semwise_credit          #get SPI using the formula
        SPI_obtained=round(SPI_obtained,2)                    #accuracy upto 2 decimal places
        CPI_obtained=total_obtained/total_credit              #get CPI using the formula
        CPI_obtained=round(CPI_obtained,2)                    #accuracy upto 2 decimal places
        sheetname[column_letter+str(5)]=semwise_credit   #fill value of total credits in a given semester
        sheetname[column_letter+str(6)]=SPI_obtained              #fill value of SPI
        sheetname[column_letter+str(7)]=total_credit     #fill value of total credits upto given semester
        sheetname[column_letter+str(8)]=CPI_obtained              #fill value of CPI

    wb.save(r'output\\'+'%s.xlsx'%roll_no)         #save the xlsx file with roll number as its name
    return


name_dict={}       # create empty dictionary "roll number" : "name"
LTP_dict={}         # create empty dictionary "subject no" : "L-T-P"
sub_name={}        # create empty dictionary "subject no" : "subject name"

with open('names-roll.csv') as f:         #opening csv file
    reader = csv.reader(f)
    i=1
    for line in reader:
        if i==1:                         #ignoring the first line as it is heading
            i=0
            continue
        name_dict[line[0]]=line[1]       #fill dictionary with 'roll_no':'name'

with open('subjects_master.csv') as f:
    reader = csv.reader(f)
    i=1
    for line in reader:
        if i==1:                         #ignoring the first line as it is heading
            i=0
            continue
        sub_name[line[0]]=line[1]         #fill dictionary with 'subject code':'subject name'
        LTP_dict[line[0]]=line[2]         #fill dictionary with 'subject code':'L-T-P'

folder = "output"   # folder
parent_folder = "C:\\Users\\narendra tejasvi\\OneDrive\\Desktop\\CS384_tut05"  # Path of parent folder
final_path = os.path.join(parent_folder, folder)   #final path after joining
os.mkdir(final_path)        # Create the folder

#dictionary to stores score for a given grade
score_dict={'F*':0, 'DD*':4, 'I':0, 'F':0, 'DD':4, 'CD':5, 'CC':6, 'BC':7, ' BB':8, 'BB':8, 'AB':9, 'AA':10 }

roll_no="0401CS01"                 #initialise it with the starting roll number
student_info=[]                    #list to store all the marks for a given roll number
file=open("grades.csv", "r")       #open csv file  in reading mode
head = file.readline()             #read the heading of the csv file

for row in file:       
    row=row.split(',')         
    if row[0]==roll_no:        
        student_info.append(row)
    else:
        # call generate_marksheetname function for current roll number
        # and then update value of current roll number
        generate_marksheetname(name_dict, sub_name, LTP_dict, student_info,  score_dict, roll_no)
        student_info=[]    #make student_info empty for new roll number
        student_info.append(row)
        roll_no=row[0]

#generate_marksheetname function for the last roll number will not be called in the loop
#it will be called outside the loop
generate_marksheetname(name_dict, sub_name, LTP_dict, student_info,  score_dict, roll_no)




