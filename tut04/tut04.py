import csv
import os
from openpyxl import Workbook,load_workbook

def output_by_subject(new_entry, heading):
    path=r"output_by_subject\\"                 #getting directory 
    if os.path.exists(path+'%s.xlsx' %new_entry[2]) == False:     #if file of this path doesn't exits already 
        wb = Workbook()                   #then create a new workbook
        sheet = wb.active                 #get the current sheet
        sheet.append(heading)             #add heading in that sheet
        wb.save(path+'%s.xlsx' %new_entry[2])                     #save the workbook

    wb = load_workbook(path+'%s.xlsx' %new_entry[2])              #load the existing workbook
    sheet = wb.active                     #get the active sheet
    sheet.append(new_entry)            #add new_entry in that sheet
    wb.save(path+'%s.xlsx' %new_entry[2])                         #save the workbook
    return

def output_individual_roll(new_entry, heading):
    path=r"output_individual_roll\\"                 #getting directory
    if os.path.exists(path+'%s.xlsx' %new_entry[0]) == False:     #if file of this path doesn't exits already 
        wb = Workbook()                   #then create a new workbook
        sheet = wb.active                 #get the current sheet
        sheet.append(heading)             #add heading in that sheet
        wb.save(path+'%s.xlsx' %new_entry[0])                     #save the workbook

    wb = load_workbook(path+'%s.xlsx' %new_entry[0])              #load the existing workbook
    sheet = wb.active                     #get the active sheet
    sheet.append(new_entry)            #add new_entry in that sheet
    wb.save(path+'%s.xlsx' %new_entry[0])                         #save the workbook
    return

subject = "output_by_subject"    #name of subject folder
first_dir = "C:\\Users\\narel\\OneDrive\\Documents\\GitHub\\1901EE41_2021\\tut04"      # Parent Directory path
first_path = os.path.join(first_dir, subject) #join path of first parent directory and name of directory
os.mkdir(first_path)    # Create the first directory

roll_no = "output_individual_roll"    #name of roll no folder
second_dir = "C:\\Users\\narel\\OneDrive\\Documents\\GitHub\\1901EE41_2021\\tut04"       # Parent Directory path
second_path = os.path.join(second_dir, roll_no) #join path of second parent directory and name of directory
os.mkdir(second_path)    # Create the second directory


with open('regtable_old.csv', 'r') as read_obj:      
    reader = csv.reader(read_obj)
    row_count = 0
    heading=[]
    for line in reader:        #iterating through the list
        if  row_count == 0:    #if line count is 0 then its header line
            heading=[line[0],line[1],line[3],line[8]]     #creating heading to be pushed onto an empty file
            row_count += 1
        else:
            new_entry=[line[0],line[1],line[3],line[8]] 
            output_individual_roll(new_entry, heading)    #implementing subtask-1
            output_by_subject(new_entry, heading)         #implementing subtask-2
